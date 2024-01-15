import openpyxl
import requests
from sales_id import sales_id
from datetime import datetime
import json
import re
import subprocess

ENTRY_NUMBER = set([
    955
])

class ColumnConstants:
    NO = 0
    NAMA = 1
    TAHUN = 2
    BULAN = 3
    AREA = 4
    NEW_USER = 5
    NEW_HOSPITAL = 6
    MONTHLY_INDIVIDUAL = 7
    MONTHLY_REGION = 8
    INCENTIVE_DISTRICT = 9
    QUARTER_INDIVIDU = 10
    QUARTER_REGION = 11
    JACKPOT = 12
    FRONT_LOADING = 13
    PRICE = 14
    KOMISI = 15
    BOOSTER_PRODUCT = 16
    REGULER = 17

    MAR_2023 = 27
    APR_2023 = 28
    MEI_2023 = 29
    JUNI_2023 = 30
    JULI_2023 = 31
    AGUSTUS_2023 = 32
    SEPTEMBER_2023 = 33
    OKTOBER_2023 = 34
    NOVEMBER_2023 = 35
    DESEMBER_2023 = 36

class ValueConstants:
    BAYAR = "Y"
    BELUM_BAYAR = "N"
    RESTRICTED = [None, "-", "", "#REF!", "RESIGN", "VACANT"]


log_file = open('log_file.txt', 'w+')

class SalesEntry:
    def __init__(self, data: dict):
        self.url = "http://rumapp.net/index.php/salesin/salesin_add"
        self.headers = {
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Accept-Language": "en-US,en;q=0.9,id;q=0.8",
            "Connection": "keep-alive",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Cookie": "ci_sessions=hj7941ik2h4uo6p0404ht7tcfh6cvmof",
            "Origin": "http://rumapp.net",
            "Referer": "http://rumapp.net/index.php/salesin",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "X-Requested-With": "XMLHttpRequest",
        }
        self.data = data

    def save_to_server(self):
        print(json.dumps(self.data, indent=4), "\n", file=log_file)
        response = requests.post(self.url, headers=self.headers, data=self.data, verify=False)
        print(response, file=log_file)
        return response
    
    def send_using_curl(self):
        form_data = "&".join([f"{key}={value}" for key, value in self.data.items()])
        curl_command = f"curl '{self.url}' -H 'Content-Type: application/x-www-form-urlencoded; charset=UTF-8' -H {json.dumps(self.headers)} --data-raw '{form_data} --compressed --insecure'"
        result = subprocess.run(curl_command, shell=True, capture_output=True, text=True)
        print(curl_command, file=log_file)
        print(result, file=log_file)
        return

def convert_data(data):
    return f"{data}" if data else "0"

def clean_name(string):
    pattern = re.compile(r'\([^)]*\)')
    return pattern.sub('', string).strip()

def construct_data(txt_bayar, user_id, tahun, bulan, newuser, newhospital, ed, product, collector, monthly_individu, monthly_region, quarter_individu, quarter_region, jackpot, front_loading, thn, mth2, price, komisi, reguler, pph21, notes, bayar):
    data = {
        "id": "",
        "txtbayar": txt_bayar,
        "nama": "",
        "user_id": user_id,
        "tahun": convert_data(tahun),
        "bulan": convert_data(bulan),
        "newuser": convert_data(newuser),
        "newhospital": convert_data(newhospital),
        "ed": convert_data(ed),
        "product": convert_data(product),
        "collector": convert_data(collector),
        "monthly_individu": convert_data(monthly_individu),
        "monthly_region": convert_data(monthly_region),
        "quarter_individu": convert_data(quarter_individu),
        "quarter_region": convert_data(quarter_region),
        "jackpot": convert_data(jackpot),
        "front_loading": convert_data(front_loading),
        "thn": convert_data(thn),
        "mth2": convert_data(mth2),
        "price": convert_data(price),
        "komisi": convert_data(komisi),
        "reguler": convert_data(reguler),
        "pph21": convert_data(pph21),
        "notes": convert_data(notes),
        "bayar": bayar
    }

    return data


def format_date(date: datetime):
    formatted_date = date.strftime("%d/%m/%Y") 
    return formatted_date

def is_valid_date(date_str):
    try:
        # Attempt to parse the string as a date with the specified format
        datetime.strptime(date_str, '%d/%m/%Y')
        return True
    except ValueError:
        return False

def format_number(number: int):
    return '{:,}'.format(number)

def get_value(number):
    if number not in ValueConstants.RESTRICTED:
        return number
    
    return 0

def value_pipeline(val):
    return format_number(round(get_value(val)))
    # return "0"


def read_excel_file(file_path):
    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        sheet_name = "rekap 22 23"
        sheet = workbook[sheet_name]

        max_row = sheet.max_row

        for row_number, row in enumerate(sheet.iter_rows(min_row=6, max_row=1283, values_only=True)):
            if len(ENTRY_NUMBER) > 0 and row[ColumnConstants.NO] not in ENTRY_NUMBER:
                # print(f"Skipping {row[ColumnConstants.NO]}....")
                continue
            reguler = row[ColumnConstants.REGULER]
            if reguler or reguler not in ValueConstants.RESTRICTED:
                print(f"Processing row number {row[ColumnConstants.NO]}")
                print(f"Processing row number {row[ColumnConstants.NO]}", file=log_file)
                try:
                    user_id = sales_id[clean_name(row[ColumnConstants.NAMA].upper())]
                    tahun = row[ColumnConstants.TAHUN]
                    bulan = row[ColumnConstants.BULAN]
                    newuser = value_pipeline(row[ColumnConstants.NEW_USER])
                    newhospital = value_pipeline(row[ColumnConstants.NEW_HOSPITAL])
                    ed = "0"
                    product = value_pipeline(row[ColumnConstants.BOOSTER_PRODUCT])
                    collector = "0"

                    # monthly_individu = (round(get_value(row[ColumnConstants.MONTHLY_INDIVIDUAL])))
                    monthly_individu = value_pipeline(row[ColumnConstants.MONTHLY_INDIVIDUAL])

                    monthly_region = value_pipeline(row[ColumnConstants.MONTHLY_REGION])
                    incentive_district = value_pipeline(row[ColumnConstants.INCENTIVE_DISTRICT])
                    # print("monthly region", monthly_region)
                    # print("incentive", incentive_district)
                    monthly_region = monthly_region if monthly_region != "0" else incentive_district
                    # print(monthly_region)
                    # print("\n")

                    quarter_individu = value_pipeline(row[ColumnConstants.QUARTER_INDIVIDU])
                    quarter_region = value_pipeline(row[ColumnConstants.QUARTER_REGION])
                    jackpot = value_pipeline(row[ColumnConstants.JACKPOT])
                    front_loading = value_pipeline(row[ColumnConstants.FRONT_LOADING])

                    price = value_pipeline(row[ColumnConstants.PRICE])
                    komisi = value_pipeline(row[ColumnConstants.KOMISI])
                    pph21 = "0"
                    reguler = format_number(round(reguler))

                    for idx, elem in enumerate(row[ColumnConstants.AGUSTUS_2023: ColumnConstants.AGUSTUS_2023 + 1]):
                        if elem not in ValueConstants.RESTRICTED:
                            thn = "2023"
                            mth2 = idx + 8 # need modification
                            notes = elem
                            bayar = ""
                            txtbayar = ""
                            if isinstance(notes, datetime):
                                notes = format_date(notes)
                                bayar = ValueConstants.BAYAR
                                txtbayar = "Y"
                            else:
                                if is_valid_date(notes):
                                    bayar = ValueConstants.BAYAR
                                    txtbayar = "Y"
                                else:
                                    bayar = ValueConstants.BELUM_BAYAR
                                    txtbayar = "N"
                            # bayar = ValueConstants.BELUM_BAYAR

                            data = construct_data(
                                txtbayar,
                                user_id,
                                tahun,
                                bulan,
                                newuser,
                                newhospital,
                                ed,
                                product,
                                collector,
                                monthly_individu,
                                monthly_region,
                                quarter_individu,
                                quarter_region,
                                jackpot,
                                front_loading,
                                thn,
                                mth2,
                                price,
                                komisi,
                                reguler,
                                pph21,
                                notes,
                                bayar
                            )
                            # data["nama"] = clean_name(row[ColumnConstants.NAMA])
                            # print(json.dumps(data, indent=4))
                            entry = SalesEntry(data)
                            entry.save_to_server()
                            # entry.send_using_curl()
                            # break
                except Exception as e:
                    print(f"Error when processing entry no: {row[ColumnConstants.NO]}")
                    print(f"Error when processing entry no: {row[ColumnConstants.NO]}, trace: {e}\n\n", file=log_file)


    except Exception as e:
        print(f"An error occurred in row {row_number}")
        print(f"An error occurred in row {row_number}: {e}", file=log_file)


penghambat = input("Have you modified month and index?")

excel_file_path = 'REKAP INSENTIF 2022 2023 (raw).xlsx'
read_excel_file(excel_file_path)